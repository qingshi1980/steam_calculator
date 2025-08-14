import tkinter as tk
from tkinter import ttk, filedialog
from iapws import IAPWS97
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

class WaterPropertiesGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("水和蒸汽性质计算器")
        self.root.geometry("1500x700")

        # 创建主框架和选项卡
        main_frame = ttk.Frame(root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill="both", expand=True)

        # 创建两个选项卡
        self.main_calc_tab = ttk.Frame(notebook, padding="10")
        self.saturation_query_tab = ttk.Frame(notebook, padding="10")

        notebook.add(self.main_calc_tab, text="两点物性计算")
        notebook.add(self.saturation_query_tab, text="饱和水性质查询")

        # 初始化两个选项卡的内容
        self.create_main_calculator_tab()
        self.create_saturation_query_tab()

        # 状态栏
        self.status_var = tk.StringVar()
        self.status_bar = ttk.Label(root, textvariable=self.status_var, relief="sunken")
        self.status_bar.pack(fill="x", padx=10, pady=5)

    def create_main_calculator_tab(self):
        # 创建输入框架
        input_frame = ttk.LabelFrame(self.main_calc_tab, text="输入参数", padding="10")
        input_frame.pack(fill="x", padx=10, pady=5)
        
        # 状态1输入
        state1_frame = ttk.LabelFrame(input_frame, text="状态1", padding="5")
        state1_frame.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        
        ttk.Label(state1_frame, text="温度 (°C):").grid(row=0, column=0, padx=5, pady=5)
        self.temp1_var = tk.StringVar(value="25")
        self.temp1_entry = ttk.Entry(state1_frame, textvariable=self.temp1_var)
        self.temp1_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(state1_frame, text="压力 (MPa):").grid(row=0, column=2, padx=5, pady=5)
        self.press1_var = tk.StringVar(value="0.1")
        self.press1_entry = ttk.Entry(state1_frame, textvariable=self.press1_var)
        self.press1_entry.grid(row=0, column=3, padx=5, pady=5)
        ttk.Label(state1_frame, text="(绝对压力)").grid(row=1, column=2, columnspan=2, padx=5)
        
        # 状态2输入
        state2_frame = ttk.LabelFrame(input_frame, text="状态2", padding="5")
        state2_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        ttk.Label(state2_frame, text="温度 (°C):").grid(row=0, column=0, padx=5, pady=5)
        self.temp2_var = tk.StringVar(value="100")
        self.temp2_entry = ttk.Entry(state2_frame, textvariable=self.temp2_var)
        self.temp2_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(state2_frame, text="压力 (MPa):").grid(row=0, column=2, padx=5, pady=5)
        self.press2_var = tk.StringVar(value="0.1")
        self.press2_entry = ttk.Entry(state2_frame, textvariable=self.press2_var)
        self.press2_entry.grid(row=0, column=3, padx=5, pady=5)
        ttk.Label(state2_frame, text="(绝对压力)").grid(row=1, column=2, columnspan=2, padx=5)
        
        # 质量输入
        mass_frame = ttk.LabelFrame(input_frame, text="质量流量", padding="5")
        mass_frame.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        
        ttk.Label(mass_frame, text="质量流量 (kg/h):").grid(row=0, column=0, padx=5, pady=5)
        self.mass_var = tk.StringVar(value="1000")
        self.mass_entry = ttk.Entry(mass_frame, textvariable=self.mass_var)
        self.mass_entry.grid(row=0, column=1, padx=5, pady=5)
        
        # 计算和保存按钮框架
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=0, column=3, padx=5, pady=5)
        
        # 计算按钮
        ttk.Button(button_frame, text="计算", command=self.calculate).pack(pady=5)
        # 保存按钮
        ttk.Button(button_frame, text="保存为TXT", command=self.save_txt).pack(pady=5)
        ttk.Button(button_frame, text="保存为Excel", command=self.save_excel).pack(pady=5)
        
        # 创建结果显示区域
        result_frame = ttk.LabelFrame(self.main_calc_tab, text="计算结果", padding="10")
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 创建两个表格
        tables_frame = ttk.Frame(result_frame)
        tables_frame.pack(fill="both", expand=True)
        
        # 状态1表格
        state1_result = ttk.LabelFrame(tables_frame, text="状态1性质", padding="5")
        state1_result.pack(side="left", fill="both", expand=True, padx=5)
        
        columns = ("属性", "数值", "单位")
        self.tree1 = ttk.Treeview(state1_result, columns=columns, show="headings", height=8)
        for col in columns:
            self.tree1.heading(col, text=col)
            self.tree1.column(col, width=120)
        self.tree1.pack(fill="both", expand=True)
        
        # 状态2表格
        state2_result = ttk.LabelFrame(tables_frame, text="状态2性质", padding="5")
        state2_result.pack(side="left", fill="both", expand=True, padx=5)
        
        self.tree2 = ttk.Treeview(state2_result, columns=columns, show="headings", height=8)
        for col in columns:
            self.tree2.heading(col, text=col)
            self.tree2.column(col, width=120)
        self.tree2.pack(fill="both", expand=True)
        
        # 焓差计算结果显示
        delta_frame = ttk.LabelFrame(result_frame, text="热负荷计算", padding="5")
        delta_frame.pack(fill="x", padx=5, pady=5)
        
        self.delta_text = tk.Text(delta_frame, height=4, wrap="word")
        self.delta_text.pack(fill="x", padx=5, pady=5)

    def create_saturation_query_tab(self):
        # 输入框架
        query_input_frame = ttk.LabelFrame(self.saturation_query_tab, text="查询参数", padding="10")
        query_input_frame.pack(fill="x", padx=10, pady=5)

        self.query_option_var = tk.StringVar(value="temp")
        
        # 查询选项
        temp_radio = ttk.Radiobutton(query_input_frame, text="按温度查询 (°C)", variable=self.query_option_var, value="temp")
        temp_radio.grid(row=0, column=0, padx=5, pady=5)
        self.query_temp_var = tk.StringVar(value="100")
        self.query_temp_entry = ttk.Entry(query_input_frame, textvariable=self.query_temp_var)
        self.query_temp_entry.grid(row=0, column=1, padx=5, pady=5)

        press_radio = ttk.Radiobutton(query_input_frame, text="按压力查询 (MPa)", variable=self.query_option_var, value="press")
        press_radio.grid(row=1, column=0, padx=5, pady=5)
        self.query_press_var = tk.StringVar(value="0.101325")
        self.query_press_entry = ttk.Entry(query_input_frame, textvariable=self.query_press_var)
        self.query_press_entry.grid(row=1, column=1, padx=5, pady=5)

        # 查询按钮
        query_button = ttk.Button(query_input_frame, text="查询", command=self.query_saturation_properties)
        query_button.grid(row=0, column=2, rowspan=2, padx=20, pady=5)

        # 结果框架
        query_result_frame = ttk.LabelFrame(self.saturation_query_tab, text="饱和性质结果", padding="10")
        query_result_frame.pack(fill="both", expand=True, padx=10, pady=5)

        columns = ("属性", "饱和水", "饱和蒸汽", "单位")
        self.saturation_tree = ttk.Treeview(query_result_frame, columns=columns, show="headings", height=15)
        for col in columns:
            self.saturation_tree.heading(col, text=col)
            self.saturation_tree.column(col, width=150, anchor="center")
        self.saturation_tree.pack(fill="both", expand=True)

    def query_saturation_properties(self):
        try:
            # 清空旧数据
            for item in self.saturation_tree.get_children():
                self.saturation_tree.delete(item)

            option = self.query_option_var.get()
            
            if option == "temp":
                temp_c = float(self.query_temp_var.get())
                temp_k = temp_c + 273.15
                sat_liquid = IAPWS97(T=temp_k, x=0)
                sat_vapor = IAPWS97(T=temp_k, x=1)
            else: # press
                press_mpa = float(self.query_press_var.get())
                sat_liquid = IAPWS97(P=press_mpa, x=0)
                sat_vapor = IAPWS97(P=press_mpa, x=1)

            properties = [
                ("温度", "T", "°C"),
                ("压力", "P", "MPa"),
                ("密度", "rho", "kg/m³"),
                ("比焓", "h", "kJ/kg"),
                ("比熵", "s", "kJ/kg·K"),
                ("比内能", "u", "kJ/kg"),
                ("比容", "v", "m³/kg"),
                ("定压比热", "cp", "kJ/kg·K"),
                ("定容比热", "cv", "kJ/kg·K"),
                ("粘度", "mu", "Pa·s"),
                ("导热系数", "k", "W/m·K")
            ]
            
            # 汽化潜热单独计算
            latent_heat = sat_vapor.h - sat_liquid.h

            for prop, attr, unit in properties:
                if attr == "T" and unit == "°C":
                    val_liq = sat_liquid.T - 273.15
                    val_vap = sat_vapor.T - 273.15
                else:
                    val_liq = getattr(sat_liquid, attr)
                    val_vap = getattr(sat_vapor, attr)
                
                self.saturation_tree.insert("", "end", values=(prop, f"{val_liq:.4f}", f"{val_vap:.4f}", unit))
            
            self.saturation_tree.insert("", "end", values=("汽化潜热", "-", f"{latent_heat:.4f}", "kJ/kg"))

            self.status_var.set("饱和性质查询成功！")

        except Exception as e:
            self.status_var.set(f"查询错误: {str(e)}")

    def save_txt(self):
        try:
            # 获取当前时间作为默认文件名
            default_filename = f"水蒸汽计算_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            
            # 打开文件保存对话框
            file_path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                initialfile=default_filename,
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
            )
            
            if not file_path:  # 如果用户取消保存
                return
            
            # 保存txt文件
            with open(file_path, 'w', encoding='utf-8') as f:
                # 写入标题
                f.write("水和蒸汽性质计算报告\n")
                f.write("=" * 50 + "\n\n")
                
                # 写入输入参数
                f.write("输入参数：\n")
                f.write("-" * 30 + "\n")
                f.write(f"状态1 - 温度: {self.temp1_var.get()} °C, 压力: {self.press1_var.get()} MPa\n")
                f.write(f"状态2 - 温度: {self.temp2_var.get()} °C, 压力: {self.press2_var.get()} MPa\n")
                f.write(f"质量流量: {self.mass_var.get()} kg/h\n\n")
                
                # 写入状态1结果
                f.write("状态1性质：\n")
                f.write("-" * 30 + "\n")
                for item in self.tree1.get_children():
                    values = self.tree1.item(item)['values']
                    f.write(f"{values[0]}: {values[1]} {values[2]}\n")
                f.write("\n")
                
                # 写入状态2结果
                f.write("状态2性质：\n")
                f.write("-" * 30 + "\n")
                for item in self.tree2.get_children():
                    values = self.tree2.item(item)['values']
                    f.write(f"{values[0]}: {values[1]} {values[2]}\n")
                f.write("\n")
                
                # 写入热负荷计算结果
                f.write("热负荷计算结果：\n")
                f.write("-" * 30 + "\n")
                f.write(self.delta_text.get(1.0, tk.END))
                
                # 写入保存时间
                f.write("\n" + "=" * 50 + "\n")
                f.write(f"计算时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
            self.status_var.set(f"结果已保存至: {file_path}")
        except Exception as e:
            self.status_var.set(f"保存错误：{str(e)}")

    def save_excel(self):
        try:
            # 获取当前时间作为默认文件名
            default_filename = f"水蒸汽计算_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # 打开文件保存对话框
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_filename,
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if not file_path:  # 如果用户取消保存
                return
            
            # 保存Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "水和蒸汽性质计算"
            row = 1
            # 输入参数
            ws.cell(row=row, column=1, value="输入参数")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
            ws.cell(row=row, column=1, value="状态1 - 温度")
            ws.cell(row=row, column=2, value=self.temp1_var.get())
            ws.cell(row=row, column=3, value="°C")
            ws.cell(row=row, column=4, value="状态1 - 压力")
            ws.cell(row=row, column=5, value=self.press1_var.get())
            ws.cell(row=row, column=6, value="MPa")
            row += 1
            ws.cell(row=row, column=1, value="状态2 - 温度")
            ws.cell(row=row, column=2, value=self.temp2_var.get())
            ws.cell(row=row, column=3, value="°C")
            ws.cell(row=row, column=4, value="状态2 - 压力")
            ws.cell(row=row, column=5, value=self.press2_var.get())
            ws.cell(row=row, column=6, value="MPa")
            row += 1
            ws.cell(row=row, column=1, value="质量流量")
            ws.cell(row=row, column=2, value=self.mass_var.get())
            ws.cell(row=row, column=3, value="kg/h")
            row += 2
            # 状态1性质
            ws.cell(row=row, column=1, value="状态1性质")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
            ws.append(["属性", "数值", "单位"])
            for item in self.tree1.get_children():
                values = self.tree1.item(item)['values']
                ws.append(values)
            row = ws.max_row + 2
            # 状态2性质
            ws.cell(row=row, column=1, value="状态2性质")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
            ws.append(["属性", "数值", "单位"])
            for item in self.tree2.get_children():
                values = self.tree2.item(item)['values']
                ws.append(values)
            row = ws.max_row + 2
            # 热负荷计算结果
            ws.cell(row=row, column=1, value="热负荷计算结果")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            row += 1
            for line in self.delta_text.get(1.0, tk.END).strip().split('\n'):
                ws.cell(row=row, column=1, value=line)
                row += 1
            # 美化Excel
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1 or cell.row in [ws.max_row - 1, ws.max_row - 2]:
                        cell.fill = header_fill
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            # 保存Excel
            wb.save(file_path)
            
            self.status_var.set(f"Excel已保存至: {file_path}")
        except Exception as e:
            self.status_var.set(f"保存错误：{str(e)}")

    def calculate(self):
        try:
            # 清空现有结果
            for item in self.tree1.get_children():
                self.tree1.delete(item)
            for item in self.tree2.get_children():
                self.tree2.delete(item)
            self.delta_text.delete(1.0, tk.END)
            
            # 获取输入值并转换温度单位
            temp1 = float(self.temp1_var.get()) + 273.15  # 摄氏度转开尔文
            press1 = float(self.press1_var.get())
            temp2 = float(self.temp2_var.get()) + 273.15  # 摄氏度转开尔文
            press2 = float(self.press2_var.get())
            mass_flow = float(self.mass_var.get())  # kg/h
            
            # 计算两个状态的水性质
            water1 = IAPWS97(T=temp1, P=press1)
            water2 = IAPWS97(T=temp2, P=press2)
            
            # 定义要显示的属性
            properties = [
                ("温度", "T", "K"),
                ("温度", "T", "°C"),
                ("压力", "P", "MPa"),
                ("密度", "rho", "kg/m³"),
                ("比焓", "h", "kJ/kg"),
                ("比熵", "s", "kJ/kg·K"),
                ("比内能", "u", "kJ/kg"),
                ("比容", "v", "m³/kg"),
                ("干度", "x", "-")
            ]
            
            # 将结果添加到表格中
            for prop, attr, unit in properties:
                if attr == "T" and unit == "°C":
                    value1 = water1.T - 273.15
                    value2 = water2.T - 273.15
                else:
                    value1 = getattr(water1, attr) if hasattr(water1, attr) else "单相"
                    value2 = getattr(water2, attr) if hasattr(water2, attr) else "单相"
                
                self.tree1.insert("", "end", values=(prop, f"{value1:.4f}" if isinstance(value1, (int, float)) else value1, unit))
                self.tree2.insert("", "end", values=(prop, f"{value2:.4f}" if isinstance(value2, (int, float)) else value2, unit))
            
            # 计算焓差和热负荷
            delta_h = water2.h - water1.h  # kJ/kg
            # 将kg/h转换为kg/s，然后计算热负荷（MW）
            mass_flow_kg_s = mass_flow / 3600  # 转换为kg/s
            heat_load = delta_h * mass_flow_kg_s / 1000  # 转换为MW
            
            # 显示计算结果
            self.delta_text.insert(tk.END, f"比焓差: {delta_h:.2f} kJ/kg\n")
            self.delta_text.insert(tk.END, f"质量流量: {mass_flow:.2f} kg/h\n")
            self.delta_text.insert(tk.END, f"热负荷: {heat_load:.3f} MW\n")
            self.delta_text.insert(tk.END, f"热负荷: {heat_load*1000:.2f} kW")
            
            self.status_var.set("计算成功！")
            
        except Exception as e:
            self.status_var.set(f"错误：{str(e)}")

def main():
    root = tk.Tk()
    app = WaterPropertiesGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
